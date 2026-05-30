import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

class Excel:
    @staticmethod
    def salvar_excel(temp: list, numeronota: int, nomeCliente: str, CPFformatado: str):
        
        if not os.path.exists('./Notas'):
            os.makedirs('./Notas')
            
        relatorio = pd.DataFrame(temp, columns=['Número', 'Ordem', 'Item', 'Quantidade', 'Valor'])
        
       
        pd.options.display.float_format = '{:.2f}'.format
        
        print("\n--- Visualização dos Dados ---")
        print(relatorio.to_string(index=False))
        
        caminho_arquivo = f'./Notas/Nº {numeronota}.xlsx'
        
        relatorio.to_excel(caminho_arquivo, index=False)
        
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        
        ws.insert_rows(1, amount=6)
        
        ws.merge_cells('A1:E2')
        titulo = ws['A1']
        titulo.value = f"NOTA FISCAL DE SERVIÇOS/PRODUTOS - Nº {numeronota}"
        titulo.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        titulo.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") 
        
        # Dados da Empresa 
        ws['A3'] = "Empresa: Uhalace de Souza S/A"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = "CNPJ: 00.000.000/0001-00"

        # Dados do Cliente 
        ws['A5'] = f"Cliente: {nomeCliente}"
        ws['A5'].font = Font(bold=True)
        ws['A6'] = f"CPF: {CPFformatado}"
        
        # A linha de cabeçalho da tabela
        for col in range(1, 6):
            celula_cabecalho = ws.cell(row=7, column=col)
            celula_cabecalho.font = Font(bold=True)
            celula_cabecalho.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['E'].width = 15 
        
      
        for linha in range(8, ws.max_row + 1):
            celula_valor = ws.cell(row=linha, column=5) 
            celula_valor.number_format = '#,##0.00'    
            celula_quantidade = ws.cell(row=linha, column=4)
            celula_quantidade.number_format = '#,##0.00'
        
        linha_final = ws.max_row + 1
        ws.merge_cells(start_row=linha_final, start_column=1, end_row=linha_final, end_column=4)
        
        texto_total = ws.cell(row=linha_final, column=1)
        texto_total.value = "VALOR TOTAL DA NOTA:"
        texto_total.font = Font(bold=True)
        texto_total.alignment = Alignment(horizontal='right') 
        
        soma_total = (relatorio['Quantidade'] * relatorio['Valor']).sum()
        
        celula_valor_total = ws.cell(row=linha_final, column=5)
        celula_valor_total.value = f"R$ {soma_total:.2f}"
        celula_valor_total.font = Font(bold=True, color="FF0000") 
        
        wb.save(caminho_arquivo)
        print(f"\nNota Fiscal formatada e salva com sucesso em: {caminho_arquivo}")